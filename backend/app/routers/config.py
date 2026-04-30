"""Excel-based bulk import/export for Configuration data.

IDs are NEVER user-provided. All matching uses natural business keys:
- Teams: name (case-insensitive)
- Projects: code (case-sensitive, conventionally upper)
- Persons: (employee_id, team) — a person who belongs to multiple teams
  has multiple rows sharing the same employee_id, one per team.
  Rows without employee_id are always created (cannot be updated via import).
- SubProjects: (project_code, name) — same name allowed across different projects.

To "delete" a record, set its `active` column to FALSE; rows are never removed
by import. To rename a Team (`name`) or Project (`code`) — use the UI.
"""

from __future__ import annotations

import io
from decimal import Decimal, InvalidOperation
from typing import Any

from fastapi import APIRouter, Depends, HTTPException, UploadFile, File
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic import BaseModel
from sqlalchemy.orm import Session

from app import models
from app.core.db import get_db

router = APIRouter(prefix="/config", tags=["config"])


SHEETS: dict[str, list[str]] = {
    "Teams": ["name", "description", "manager", "active"],
    "Persons": [
        "employee_id", "name", "email", "location", "line_manager",
        "allocation", "employment_type", "funding", "team", "active",
    ],
    "Projects": ["code", "name", "description", "funding", "active"],
    "SubProjects": ["project_code", "name", "description", "funding", "active"],
}

EMPLOYMENT_TYPES = {"Permanent", "Contractor", "Intern"}

HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", fgColor="DB0011")


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        ws.column_dimensions[get_column_letter(col)].width = max(14, len(h) + 4)
    ws.freeze_panes = "A2"


def _bool_str(v: bool) -> str:
    return "TRUE" if v else "FALSE"


def _build_workbook(db: Session, *, with_data: bool) -> Workbook:
    wb = Workbook()
    default = wb.active
    if default is not None:
        wb.remove(default)

    ws = wb.create_sheet("Teams")
    _write_header(ws, SHEETS["Teams"])
    if with_data:
        for t in db.query(models.Team).order_by(models.Team.name).all():
            ws.append([t.name, t.description or "", t.manager or "", _bool_str(t.active)])

    ws = wb.create_sheet("Persons")
    _write_header(ws, SHEETS["Persons"])
    if with_data:
        rows = (
            db.query(models.Person, models.Team)
            .join(models.Team, models.Team.id == models.Person.team_id)
            .order_by(models.Person.name, models.Team.name)
            .all()
        )
        for p, t in rows:
            ws.append([
                p.employee_id or "", p.name, p.email or "",
                p.location or "", p.line_manager or "",
                float(p.allocation) if p.allocation is not None else "",
                p.employment_type or "",
                p.funding or "",
                t.name, _bool_str(p.active),
            ])

    ws = wb.create_sheet("Projects")
    _write_header(ws, SHEETS["Projects"])
    if with_data:
        for pr in db.query(models.Project).order_by(models.Project.code).all():
            ws.append([
                pr.code, pr.name, pr.description or "",
                pr.funding or "", _bool_str(pr.active),
            ])

    ws = wb.create_sheet("SubProjects")
    _write_header(ws, SHEETS["SubProjects"])
    if with_data:
        rows = (
            db.query(models.SubProject, models.Project)
            .join(models.Project, models.SubProject.project_id == models.Project.id)
            .order_by(models.Project.code, models.SubProject.name)
            .all()
        )
        for sp, pr in rows:
            ws.append([
                pr.code, sp.name,
                sp.description or "", sp.funding or "",
                _bool_str(sp.active),
            ])

    info = wb.create_sheet("Instructions", 0)
    info.append(["Book of Work — Configuration Template"])
    info["A1"].font = Font(bold=True, size=14)
    info.append([])
    info.append(["IDs are NEVER user-provided. The system matches existing rows by natural keys:"])
    info.append(["  • Teams: matched by `name` (case-insensitive)."])
    info.append(["  • Projects: matched by `code`."])
    info.append(["  • Persons: matched by `(employee_id, team)`. Rows without employee_id are always created as new."])
    info.append(["  • SubProjects: matched by `(project_code, name)`."])
    info.append([])
    info.append(["• Rows are never deleted by import — set `active` to FALSE to retire."])
    info.append(["• To rename a Team (`name`) or a Project (`code`), use the Configuration UI — editing them here would create a new row instead."])
    info.append(["• Persons.team: ONE team name per row. A person who belongs to multiple teams must have one row per team (employee_id repeats)."])
    info.append(["• Projects do NOT have a teams column — every team can pick from any project."])
    info.append(["• Persons.allocation: number 0–100 (FTE percent for that team)."])
    info.append(["• Persons.employment_type: free text (e.g. Permanent, Contractor, Intern, or any custom value)."])
    info.append(["• SubProjects.project_code: must reference an existing Projects.code (or one being created in this same upload)."])
    info.append(["• Booleans accept TRUE/FALSE/1/0 (case-insensitive)."])
    info.column_dimensions["A"].width = 110

    return wb


@router.get("/template", summary="Download configuration data as Excel")
def download_template(db: Session = Depends(get_db)) -> StreamingResponse:
    wb = _build_workbook(db, with_data=True)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="bow-config.xlsx"'},
    )


def _parse_bool(v: Any) -> bool | None:
    if v is None or v == "":
        return None
    if isinstance(v, bool):
        return v
    s = str(v).strip().lower()
    if s in ("true", "1", "yes", "y"):
        return True
    if s in ("false", "0", "no", "n"):
        return False
    return None


def _parse_decimal(v: Any) -> Decimal | None:
    if v is None or v == "":
        return None
    try:
        return Decimal(str(v))
    except (InvalidOperation, ValueError):
        return None


class ImportError_(BaseModel):
    sheet: str
    row: int
    message: str


class ImportSummary(BaseModel):
    teams_created: int = 0
    teams_updated: int = 0
    persons_created: int = 0
    persons_updated: int = 0
    projects_created: int = 0
    projects_updated: int = 0
    sub_projects_created: int = 0
    sub_projects_updated: int = 0


class ImportResult(BaseModel):
    ok: bool
    errors: list[ImportError_] = []
    summary: ImportSummary = ImportSummary()


def _read_sheet(wb, name: str) -> list[dict[str, Any]]:
    if name not in wb.sheetnames:
        return []
    ws = wb[name]
    headers = [
        (c.value or "").strip() if isinstance(c.value, str) else c.value
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    rows: list[dict[str, Any]] = []
    for r_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in row):
            continue
        rec: dict[str, Any] = {"_row": r_idx}
        for h, v in zip(headers, row):
            if h is None:
                continue
            if isinstance(v, str):
                v = v.strip()
            rec[h] = v if v != "" else None
        rows.append(rec)
    return rows


@router.post("/import", response_model=ImportResult, summary="Import configuration from Excel")
async def import_template(
    file: UploadFile = File(...),
    dry_run: bool = False,
    db: Session = Depends(get_db),
) -> ImportResult:
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xlsm")):
        raise HTTPException(400, "Please upload a .xlsx file")

    content = await file.read()
    try:
        wb = load_workbook(io.BytesIO(content), data_only=True)
    except Exception as exc:
        raise HTTPException(400, f"Could not open workbook: {exc}") from exc

    errors: list[ImportError_] = []
    summary = ImportSummary()

    # ---------------- Teams ----------------
    team_by_name_existing: dict[str, models.Team] = {
        t.name.lower(): t for t in db.query(models.Team).all()
    }
    team_rows = _read_sheet(wb, "Teams")
    pending_teams: list[tuple[dict[str, Any], models.Team | None]] = []
    seen_team_names: set[str] = set()
    for rec in team_rows:
        row_no = rec["_row"]
        name = rec.get("name")
        if not name:
            errors.append(ImportError_(sheet="Teams", row=row_no, message="`name` is required"))
            continue
        nlow = name.lower()
        if nlow in seen_team_names:
            errors.append(ImportError_(sheet="Teams", row=row_no, message=f"duplicate team name {name!r} in sheet"))
            continue
        seen_team_names.add(nlow)
        pending_teams.append((rec, team_by_name_existing.get(nlow)))

    pending_team_names = {(r.get("name") or "").lower() for r, ex in pending_teams if not ex}

    # ---------------- Projects ----------------
    project_by_code_existing: dict[str, models.Project] = {
        p.code: p for p in db.query(models.Project).all()
    }
    project_rows = _read_sheet(wb, "Projects")
    pending_projects: list[tuple[dict[str, Any], models.Project | None]] = []
    seen_codes: set[str] = set()
    for rec in project_rows:
        row_no = rec["_row"]
        code = rec.get("code")
        name = rec.get("name")
        if not code or not name:
            errors.append(ImportError_(sheet="Projects", row=row_no, message="`code` and `name` are required"))
            continue
        if code in seen_codes:
            errors.append(ImportError_(sheet="Projects", row=row_no, message=f"duplicate code {code!r} in sheet"))
            continue
        seen_codes.add(code)
        pending_projects.append((rec, project_by_code_existing.get(code)))

    pending_project_codes = {(r.get("code") or "") for r, ex in pending_projects if not ex}

    # ---------------- Persons ----------------
    # Existing rows indexed by (employee_id, team_id) — only those WITH employee_id are matchable.
    existing_persons_by_emp_team: dict[tuple[str, str], models.Person] = {}
    for p in db.query(models.Person).filter(models.Person.employee_id.isnot(None)).all():
        existing_persons_by_emp_team[(p.employee_id, p.team_id)] = p  # type: ignore[index]

    person_rows = _read_sheet(wb, "Persons")
    pending_persons: list[tuple[dict[str, Any], models.Person | None, str]] = []
    seen_emp_team_in_sheet: set[tuple[str, str]] = set()
    for rec in person_rows:
        row_no = rec["_row"]
        name = rec.get("name")
        team_name = rec.get("team")
        if not name:
            errors.append(ImportError_(sheet="Persons", row=row_no, message="`name` is required"))
            continue
        if not team_name:
            errors.append(ImportError_(sheet="Persons", row=row_no, message="`team` is required (one team per row)"))
            continue
        tlow = team_name.lower()
        if tlow not in team_by_name_existing and tlow not in pending_team_names:
            errors.append(ImportError_(sheet="Persons", row=row_no, message=f"unknown team {team_name!r}"))
            continue

        emp_id = rec.get("employee_id")
        if emp_id:
            key = (str(emp_id), tlow)
            if key in seen_emp_team_in_sheet:
                errors.append(ImportError_(sheet="Persons", row=row_no, message=f"duplicate row for employee_id {emp_id!r} in team {team_name!r}"))
                continue
            seen_emp_team_in_sheet.add(key)

        emp_type = rec.get("employment_type")
        if emp_type and len(str(emp_type)) > 20:
            errors.append(ImportError_(sheet="Persons", row=row_no, message="employment_type must be at most 20 characters"))
            continue

        if rec.get("allocation") is not None:
            alloc = _parse_decimal(rec.get("allocation"))
            if alloc is None or alloc < 0 or alloc > 100:
                errors.append(ImportError_(sheet="Persons", row=row_no, message="allocation must be a number between 0 and 100"))
                continue

        # Match existing only when emp_id is provided AND team already exists
        person_existing: models.Person | None = None
        if emp_id and tlow in team_by_name_existing:
            existing_team = team_by_name_existing[tlow]
            person_existing = existing_persons_by_emp_team.get((str(emp_id), existing_team.id))
        pending_persons.append((rec, person_existing, team_name))

    # ---------------- SubProjects ----------------
    # Existing rows indexed by (project_code, name.lower())
    existing_subs: dict[tuple[str, str], models.SubProject] = {}
    for sp, pr in (
        db.query(models.SubProject, models.Project)
        .join(models.Project, models.Project.id == models.SubProject.project_id)
        .all()
    ):
        existing_subs[(pr.code, sp.name.lower())] = sp

    sp_rows = _read_sheet(wb, "SubProjects")
    pending_sps: list[tuple[dict[str, Any], models.SubProject | None]] = []
    seen_sp_keys: set[tuple[str, str]] = set()
    for rec in sp_rows:
        row_no = rec["_row"]
        name = rec.get("name")
        pcode = rec.get("project_code")
        if not name or not pcode:
            errors.append(ImportError_(sheet="SubProjects", row=row_no, message="`name` and `project_code` are required"))
            continue
        if pcode not in project_by_code_existing and pcode not in pending_project_codes:
            errors.append(ImportError_(sheet="SubProjects", row=row_no, message=f"unknown project_code {pcode!r}"))
            continue
        key = (pcode, name.lower())
        if key in seen_sp_keys:
            errors.append(ImportError_(sheet="SubProjects", row=row_no, message=f"duplicate sub-project {name!r} under project {pcode!r} in sheet"))
            continue
        seen_sp_keys.add(key)
        pending_sps.append((rec, existing_subs.get(key)))

    if errors:
        return ImportResult(ok=False, errors=errors, summary=summary)

    if dry_run:
        return ImportResult(ok=True, errors=[], summary=summary)

    # ---------------- Apply ----------------
    team_by_name: dict[str, models.Team] = dict(team_by_name_existing)

    for rec, team_existing in pending_teams:
        active = _parse_bool(rec.get("active"))
        if team_existing:
            team_existing.description = rec.get("description")
            team_existing.manager = rec.get("manager")
            if active is not None:
                team_existing.active = active
            summary.teams_updated += 1
        else:
            t = models.Team(
                name=rec["name"],
                description=rec.get("description"),
                manager=rec.get("manager"),
                active=active if active is not None else True,
            )
            db.add(t)
            team_by_name[t.name.lower()] = t
            summary.teams_created += 1
    db.flush()

    project_by_code: dict[str, models.Project] = dict(project_by_code_existing)
    for rec, proj_existing in pending_projects:
        active = _parse_bool(rec.get("active"))
        if proj_existing:
            proj_existing.name = rec["name"]
            proj_existing.description = rec.get("description")
            proj_existing.funding = rec.get("funding")
            if active is not None:
                proj_existing.active = active
            summary.projects_updated += 1
        else:
            p = models.Project(
                code=rec["code"],
                name=rec["name"],
                description=rec.get("description"),
                funding=rec.get("funding"),
                active=active if active is not None else True,
            )
            db.add(p)
            project_by_code[p.code] = p
            summary.projects_created += 1
    db.flush()

    for rec, person_existing, team_name in pending_persons:
        active = _parse_bool(rec.get("active"))
        team = team_by_name[team_name.lower()]
        alloc = _parse_decimal(rec.get("allocation"))
        emp_type = rec.get("employment_type") or "Permanent"
        fields = dict(
            employee_id=rec.get("employee_id"),
            name=rec["name"],
            email=rec.get("email"),
            location=rec.get("location"),
            line_manager=rec.get("line_manager"),
            employment_type=emp_type,
            funding=rec.get("funding"),
            team_id=team.id,
        )
        if alloc is not None:
            fields["allocation"] = alloc
        if person_existing:
            for k, v in fields.items():
                setattr(person_existing, k, v)
            if active is not None:
                person_existing.active = active
            summary.persons_updated += 1
        else:
            if "allocation" not in fields:
                fields["allocation"] = Decimal("100")
            p = models.Person(
                **fields,
                active=active if active is not None else True,
            )
            db.add(p)
            summary.persons_created += 1

    for rec, sp_existing in pending_sps:
        active = _parse_bool(rec.get("active"))
        project = project_by_code[rec["project_code"]]
        if sp_existing:
            sp_existing.description = rec.get("description")
            sp_existing.funding = rec.get("funding")
            if active is not None:
                sp_existing.active = active
            summary.sub_projects_updated += 1
        else:
            sp = models.SubProject(
                project_id=project.id,
                name=rec["name"],
                description=rec.get("description"),
                funding=rec.get("funding"),
                active=active if active is not None else True,
            )
            db.add(sp)
            summary.sub_projects_created += 1

    db.commit()
    return ImportResult(ok=True, errors=[], summary=summary)